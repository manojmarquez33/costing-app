#!/usr/bin/env python3
"""
Extract all data, formulas, and structure from Excel file
"""
import openpyxl
import json
from openpyxl.utils import get_column_letter

def extract_excel_data(excel_file):
    """Extract all sheets, cells, values, and formulas from Excel file"""
    
    workbook = openpyxl.load_workbook(excel_file, data_only=False)
    
    result = {
        'filename': excel_file,
        'sheets': []
    }
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        sheet_data = {
            'name': sheet_name,
            'cells': [],
            'merged_cells': [],
            'dimensions': {
                'max_row': sheet.max_row,
                'max_column': sheet.max_column
            }
        }
        
        # Extract merged cells
        for merged_range in sheet.merged_cells.ranges:
            sheet_data['merged_cells'].append(str(merged_range))
        
        # Extract all cells with data or formulas
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                                   min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value is not None or cell.data_type == 'f':
                    cell_info = {
                        'address': cell.coordinate,
                        'row': cell.row,
                        'column': get_column_letter(cell.column),
                        'column_index': cell.column,
                        'value': str(cell.value) if cell.value is not None else '',
                        'data_type': cell.data_type,
                        'number_format': cell.number_format,
                    }
                    
                    # Extract formula if present
                    if cell.data_type == 'f':
                        cell_info['formula'] = f"={cell.value}"
                        # Get calculated value
                        try:
                            calc_wb = openpyxl.load_workbook(excel_file, data_only=True)
                            calc_sheet = calc_wb[sheet_name]
                            calc_cell = calc_sheet[cell.coordinate]
                            cell_info['calculated_value'] = str(calc_cell.value) if calc_cell.value is not None else ''
                        except:
                            cell_info['calculated_value'] = 'N/A'
                    
                    # Add styling info if relevant
                    if cell.font.bold:
                        cell_info['bold'] = True
                    if cell.fill.start_color.index != '00000000':
                        cell_info['background_color'] = str(cell.fill.start_color.index)
                    
                    sheet_data['cells'].append(cell_info)
        
        result['sheets'].append(sheet_data)
    
    return result

def save_to_files(data, base_filename='excel_data'):
    """Save extracted data to multiple formats"""
    
    # Save as JSON
    json_file = f"{base_filename}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"✓ Saved JSON: {json_file}")
    
    # Save as readable text file
    txt_file = f"{base_filename}.txt"
    with open(txt_file, 'w', encoding='utf-8') as f:
        f.write(f"Excel File: {data['filename']}\n")
        f.write("=" * 80 + "\n\n")
        
        for sheet in data['sheets']:
            f.write(f"\n{'='*80}\n")
            f.write(f"SHEET: {sheet['name']}\n")
            f.write(f"{'='*80}\n")
            f.write(f"Dimensions: {sheet['dimensions']['max_row']} rows × {sheet['dimensions']['max_column']} columns\n\n")
            
            if sheet['merged_cells']:
                f.write(f"Merged Cells: {', '.join(sheet['merged_cells'])}\n\n")
            
            # Group cells by row for better readability
            cells_by_row = {}
            for cell in sheet['cells']:
                row = cell['row']
                if row not in cells_by_row:
                    cells_by_row[row] = []
                cells_by_row[row].append(cell)
            
            # Write cells row by row
            for row_num in sorted(cells_by_row.keys()):
                f.write(f"\n--- Row {row_num} ---\n")
                for cell in sorted(cells_by_row[row_num], key=lambda x: x['column_index']):
                    f.write(f"\n  Cell {cell['address']} ({cell['column']}{cell['row']}):\n")
                    
                    if 'formula' in cell:
                        f.write(f"    Formula: {cell['formula']}\n")
                        f.write(f"    Calculated Value: {cell['calculated_value']}\n")
                    else:
                        f.write(f"    Value: {cell['value']}\n")
                    
                    if cell['data_type'] != 's':
                        f.write(f"    Type: {cell['data_type']}\n")
                    if cell['number_format'] != 'General':
                        f.write(f"    Format: {cell['number_format']}\n")
                    if cell.get('bold'):
                        f.write(f"    Style: BOLD\n")
                    if cell.get('background_color'):
                        f.write(f"    Background: {cell['background_color']}\n")
    
    print(f"✓ Saved Text: {txt_file}")
    
    # Save formulas only
    formulas_file = f"{base_filename}_formulas.txt"
    with open(formulas_file, 'w', encoding='utf-8') as f:
        f.write("ALL FORMULAS IN EXCEL FILE\n")
        f.write("=" * 80 + "\n\n")
        
        for sheet in data['sheets']:
            formulas = [cell for cell in sheet['cells'] if 'formula' in cell]
            if formulas:
                f.write(f"\n{'='*80}\n")
                f.write(f"SHEET: {sheet['name']}\n")
                f.write(f"{'='*80}\n\n")
                
                for cell in formulas:
                    f.write(f"{cell['address']}: {cell['formula']}\n")
                    f.write(f"  → Result: {cell['calculated_value']}\n\n")
    
    print(f"✓ Saved Formulas: {formulas_file}")

if __name__ == '__main__':
    excel_file = 'Pug-cost.xlsx'
    
    print(f"Extracting data from {excel_file}...")
    data = extract_excel_data(excel_file)
    
    print(f"\nFound {len(data['sheets'])} sheet(s):")
    for sheet in data['sheets']:
        formula_count = len([c for c in sheet['cells'] if 'formula' in c])
        print(f"  - {sheet['name']}: {len(sheet['cells'])} cells, {formula_count} formulas")
    
    print("\nSaving to files...")
    save_to_files(data)
    
    print("\n✓ Extraction complete!")
    print("\nGenerated files:")
    print("  1. excel_data.json - Full data in JSON format")
    print("  2. excel_data.txt - Human-readable format with all details")
    print("  3. excel_data_formulas.txt - All formulas only")
